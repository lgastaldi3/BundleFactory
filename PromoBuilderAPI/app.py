import base64
import re
from flask import Flask, request, send_file, jsonify, abort
from flask_cors import CORS
from PIL import Image, ImageFont
from io import BytesIO
from CardGenerator.product_card_generator import generate_card_basic
from DiscountPolicy.DiscountPolicyFactory import get_specific_volumes, generate_discounts, generate_quantities
import requests
from itertools import combinations
import random
import os
import itertools
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import StringIO
import json
import zipfile

ComboStoreAPI = "http://127.0.0.1:5001/get_mappings"

app = Flask(__name__)
CORS(app)

IMAGE_FOLDER = "./BundleImages"
font = ImageFont.truetype("CardGenerator/Assets/Fonts/TCCC-UnityCondensedPC-Bold.ttf", size=400)

themes = {
    'jm': ('khaki', 'black'),
    'monster': ('lime', 'black'), 
    'bang': ('purple', 'white'), 
    'sprite': ('limegreen', 'black'), 
    'fanta': ('orange', 'black'), 
    'vw': ('crimson', 'white'), 
    'sw': ('dodgerblue', 'white'), 
    'powerade': ('deepskyblue', 'black'), 
    'drp': ('firebrick', 'white'), 
    'kod': ('lightgrey', 'black'), 
    'koz': ('red', 'white'), 
    'kov': ('moccasin', 'black'), 
    'koc': ('fuchsia', 'white'), 
    'kof': ('magenta', 'white'), 
    'ko': ('red', 'white'), 
    'mm': ('yellowgreen', 'black'), 
    'gp': ('peru', 'white'), 
    'ba': ('coral', 'black'), 
    'reign': ('slategrey', 'white'), 
    'cp': ('darkgoldenrod', 'white'), 
    'fairlife': ('skyblue', 'black'), 
    'fuze': ('gold', 'black'), 
    'dunkin': ('darkorange', 'white'), 
    'seagram': ('darkgreen', 'white'), 
    'barq': ('darkred', 'white'), 
    'aha': ('plum', 'black'), 
    'dasani': ('aqua', 'black'), 
    'fresca': ('powderblue', 'black'), 
    'ch': ('darkturquoise', 'white'), 
    'nos': ('dodgerblue', 'white'), 
    'pt': ('gold', 'black'), 
    'ga': ('darkgreen', 'white'), 
    'tc': ('yellow', 'black'), 
    'inca': ('gold', 'black'), 
    'lp': ('silver', 'black'), 
    'rc': ('royalblue', 'white'), 
    'core': ('silver', 'black'), 
    'water': ('silver', 'black'), 
    'orange': ('silver', 'black')
}

dev_sets = ['meals at home', 'meals away from home', 'leisure', 'enjoyment', 
                'refreshment', 'energy', 'at home', 'work', 
                'breakfast', 'on the go', 'health', 'exercise', 
                'socializing']

### FOR PRINTER ###
def write_df_to_excel(df, file_path, sheet_name='dev_set'):
    try:
        # Load the existing workbook if it exists
        book = load_workbook(file_path)
    except FileNotFoundError:
        # If workbook does not exist, create a new one
        book = None

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a' if book else 'w') as writer:
        if book:
            
            # Remove the existing sheet with the same name
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]
        
        # Write DataFrame to the specified sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)

### FOR PRINTER ###
def write_df_to_excel_table(df, file_path="BundleFactory.xlsx", sheet_name='dev_set'):
    try:
        # Load the existing workbook if it exists
        book = load_workbook(file_path)
    except FileNotFoundError:
        # If workbook does not exist, create a new one
        book = None

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a' if book else 'w') as writer:
        if book:
            
            # Remove the existing sheet with the same name
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]
        
        # Write DataFrame to the specified sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Reload the workbook to access the added sheet
    book = load_workbook(file_path)
    
    # Ensure the sheet exists
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    
    sheet = book[sheet_name]

    # Define the range for the table (A1 to last cell)
    tab = Table(displayName=sheet_name.replace(" ", ""), ref=f"A1:{chr(64 + len(df.columns))}{len(df) + 1}")

    # Add a default style with striped rows
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, 
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    
    # Add the table to the worksheet
    sheet.add_table(tab)
    
    # Save the workbook
    book.save(file_path)

### FOR PRINTER ###
def images_to_urls(images):
    urls = {}
    for i, img in enumerate(images):
        # Create a BytesIO object
        img_io = BytesIO()
        bID, img = img

        # Save the image to the BytesIO object
        img.save(img_io, format='PNG')

        # Create a base64 string
        img_base64 = base64.b64encode(img_io.getvalue()).decode('utf-8')

        # Create a data URL
        img_url = f"data:image/png;base64,{img_base64}"

        # Add the URL to the list
        urls[bID] = img_url

    return urls

### FOR PRINTER ###
def images_to_files(images):
    # Ensure the SKUs directory exists
    if not os.path.exists('BundleImages'):
        os.makedirs('BundleImages')

    for i, img in enumerate(images):
        bID, img = img
        # Save the image to the SKUs directory
        img.save(f'BundleImages/{bID}.png', format='PNG')

    return {'status': 'Images saved successfully'}

### FOR COMBO BUILDER ###
def generate_combos(SKUs_with_discounts, dev_set):
    combinations = list(itertools.combinations(SKUs_with_discounts.index, 1))
    development_data = []

    for i, combination in enumerate(combinations, start=1):
        key = f"{dev_sets.index(dev_set):02d}T1D{i:06d}"
        rows = SKUs_with_discounts.loc[list(combination)]
        introductory_discounts = rows['introductory_discount'].tolist()
        value = (
            rows['PRODUCT_DESCRIPTION'].tolist(),
            rows['PACKAGE_GROUP'].tolist(),
            introductory_discounts,
            rows['introductory_quantity'].tolist(),
            round((sum(introductory_discounts) / len(introductory_discounts)), 1),
            3,
            dev_set,
            "development"
        )
        development_data.append([key] + list(value))

    columns = [
        'BundleID', 'Product_Description', 'Package_Group', 'Discounts',
        'Quantities', 'Average_Discount', 'Max_Quantity', 'Dev_Set', 'Type'
    ]

    development_df = pd.DataFrame(development_data, columns=columns).to_json()

    ########## VOLUME ##########
    volume_data = []

    combinations = (
        list(itertools.combinations(SKUs_with_discounts.index, 1)) +
        list(itertools.combinations(SKUs_with_discounts.index, 2)) +
        list(itertools.combinations(SKUs_with_discounts.index, 3))
    )

    for i, combination in enumerate(combinations, start=1):
        key = f"{dev_sets.index(dev_set):02d}T1B{i:06d}"
        rows = SKUs_with_discounts.loc[list(combination)]
        incremental_discounts = rows['incremental_discount'].tolist()
        value = (
            rows['PRODUCT_DESCRIPTION'].tolist(),
            rows['PACKAGE_GROUP'].tolist(),
            incremental_discounts,
            rows['incremental_quantity'].tolist(),
            round((sum(incremental_discounts) / len(incremental_discounts)), 1),
            3,
            dev_set,
            "volume"
        )
        volume_data.append([key] + list(value))

    volume_df = pd.DataFrame(volume_data, columns=columns).to_json()

    assortment_data = []
    combinations = list(itertools.combinations(SKUs_with_discounts.index, 2))
    for i, (index1, index2) in enumerate(combinations, start=1):
        (index1, row1), (index2, row2) = SKUs_with_discounts.loc[[index1, index2]].iterrows()
        if row1['VOLUME_RAW_CASES'] > row2['VOLUME_RAW_CASES']:
            key = f"{dev_sets.index(dev_set):02d}T1B{i:06d}"
            assortment_discounts = [row1['anchor_discount'], row2['incremental_discount']]
            value = (
                [row1['PRODUCT_DESCRIPTION'], row2['PRODUCT_DESCRIPTION']],
                [row1['PACKAGE_GROUP'], row2['PACKAGE_GROUP']],
                assortment_discounts,
                [row1['anchor_quantity'], row2['introductory_quantity']],
                round((sum(assortment_discounts) / len(assortment_discounts)), 1),
                3,
                dev_set,
                "assortment"
            )
            assortment_data.append([key] + list(value))
    assortment_df = pd.DataFrame(assortment_data, columns=columns).to_json()
    return {"Development Discounts" : development_df, "Assortment Bundles" : assortment_df, "Volume Bundles" : volume_df}

@app.route('/get-dev-set-names', methods=['GET'])
def get_dev_set_names():
    print("Getting Dev Set Names")
    return dev_sets

@app.route('/populate_tables', methods=['POST'])
def populate_tables():
    print("Populating Tables")
    request_data = request.get_json()
    table_dict = request_data['tables']
    dev_set = request_data['dev_set']
    parsed_json_data = {key: json.loads(value) for key, value in table_dict.items()}
    bundle_tables = {key: pd.read_json(StringIO(json.dumps(value)), orient='columns') for key, value in parsed_json_data.items()}
    bundle_table = pd.concat([item[1] for item in bundle_tables.items()]).set_index('BundleID')
    bundle_table = bundle_table.reset_index().rename(columns={'index': 'Bundle_ID'})
    file_path = 'BundleFactory.xlsx'

    write_df_to_excel_table(bundle_table, file_path, sheet_name=dev_set)
    return "SUCCESS"

@app.route('/populate-all-tables', methods=['POST'])
def populate_all_tables():
    print("Populating All Tables")
    num_items = int(request.form.get('num_items'))
    introductory_discount = int(request.form.get('introductory_discount'))
    incremental_discount = int(request.form.get('incremental_discount'))
    anchor_discount = int(request.form.get('anchor_discount'))

    generate_discounts(introductory_discount, incremental_discount, anchor_discount)
    generate_quantities()

    response = requests.request("GET", ComboStoreAPI, headers={}, data={}).json()
    for dev_set in dev_sets:
        SKUs = response['SKUs'][dev_set][:num_items]
        SKUs_with_discounts = get_specific_volumes(SKUs)
        combos = generate_combos(SKUs_with_discounts, dev_set)
        parsed_json_data = {key: json.loads(value) for key, value in combos.items()}
        bundle_tables = {key: pd.read_json(StringIO(json.dumps(value)), orient='columns') for key, value in parsed_json_data.items()}
        bundle_table = pd.concat([item[1] for item in bundle_tables.items()]).set_index('BundleID')
        bundle_table = bundle_table.reset_index().rename(columns={'index': 'Bundle_ID'})
        write_df_to_excel_table(bundle_table, sheet_name=dev_set)
        
    return "Successfully Created Tables"

@app.route('/process_image', methods=['POST'])
def process_image():
    print("Processing Single Image")
    SKUs = request.form.get('SKUs').split(',')
    dev_set = request.form.get('dev_set')
    bundle_ID = request.form.get('bundle_ID')
    quantities = request.form.get('quantities').split(',')
    max_quantity = request.form.get('max_quantity')
    discount = request.form.get('discount')
    brand = request.form.get('brand')

    # convert string data to PIL Image
    img = []
    for key in request.files.values():
        file_str = key.read()
        img.append(Image.open(BytesIO(file_str)))

    product_card = generate_card_basic(SKUs,
                                 dev_set,
                                 bundle_ID,
                                 quantities,
                                 max_quantity,
                                 font,
                                 discount,
                                 theme_color=themes[brand][0],
                                 text_color=themes[brand][1],
                                 bg_color='none')

    img_rgba = product_card.convert('RGBA')

    # Save to a BytesIO object
    img_io = BytesIO()
    img_rgba.save(img_io, 'PNG')
    img_io.seek(0)

    return send_file(img_io, mimetype='image/png')

@app.route('/generate_dev_set_bundles', methods=['POST'])
def generate_dev_set_bundles():
    print("Generating Bundles")
    dev_set = request.form.get('dev_set')
    num_items = int(request.form.get('num_items'))
    introductory_discount = int(request.form.get('introductory_discount'))
    incremental_discount = int(request.form.get('incremental_discount'))
    anchor_discount = int(request.form.get('anchor_discount'))

    response = requests.request("GET", ComboStoreAPI, headers={}, data={}).json()
    SKUs = response['SKUs'][dev_set][:num_items]

    generate_discounts(introductory_discount, incremental_discount, anchor_discount)
    generate_quantities()
    SKUs_with_discounts = get_specific_volumes(SKUs)

    return generate_combos(SKUs_with_discounts, dev_set)

@app.route('/generate_images_from_bundles', methods=['POST'])
def generate_images_from_bundles():
    print("Generating and Saving Images for Dev Set Bundles")
    response = requests.request("GET", ComboStoreAPI, headers={}, data={}).json()
    brands = response['brands']

    table_dict = request.get_json()
    parsed_json_data = {key: json.loads(value) for key, value in table_dict.items()}
    bundle_tables = {key: pd.read_json(StringIO(json.dumps(value)), orient='columns') for key, value in parsed_json_data.items()}
    bundle_table = pd.concat([item[1] for item in bundle_tables.items()]).set_index('BundleID')
    bundle_table = bundle_table.reset_index().rename(columns={'index': 'Bundle_ID'})

    product_cards = []

    for index, row in bundle_table.iterrows():
        product_description = row["Product_Description"]
        dev_set = row["Dev_Set"]
        bundle_ID = row["BundleID"]
        quantities = row["Quantities"]
        max_quantity = row["Max_Quantity"]
        avg_discount = round(row["Average_Discount"], 1)

        product_cards.append((bundle_ID, generate_card_basic(
            product_description,
            dev_set,
            bundle_ID,
            quantities,
            max_quantity,
            font,
            str(avg_discount),
            theme_color=themes[brands[product_description[0]]][0],
            text_color=themes[brands[product_description[0]]][1],
            bg_color='none'
        )))

    return images_to_files(product_cards)

@app.route('/generate_dev_set_bundle_images', methods=['POST'])
def generate_dev_set_bundle_images():
    print("Generating and Resending Images for Bundles")
    dev_set = request.form.get('dev_set')
    num_items = int(request.form.get('num_items'))

    response = requests.request("GET", ComboStoreAPI, headers={}, data={}).json()
    num_items = min(num_items, len(response['SKUs']))
    SKUs = response['SKUs'][dev_set][:num_items]

    bundles = [list(pair) for pair in list(combinations(SKUs, 1)) + list(combinations(SKUs, 2))]
    brands = response['brands']

    product_cards = []
    quantities = [[random.randint(1, 4) for product in bundle] for bundle in bundles]
    discounts = [random.randint(1, 6) for bundle in bundles]
    max_quantities = [random.randint(1, 3) for bundle in bundles]
    i = 0
    
    for bundle in bundles:
        bundle_ID = ("T1D" if len(bundle) == 1 else "T2B") + str(i)
        product_cards.append([bundle_ID, generate_card_basic(bundle,
                                 dev_set,
                                 bundle_ID,
                                 quantities[i],
                                 max_quantities[i],
                                 font,
                                 str(discounts[i]),
                                 theme_color=themes[brands[bundle[0]]][0],
                                 text_color=themes[brands[bundle[0]]][1],
                                 bg_color='none')])
        i += 1
    
    urls = images_to_urls(product_cards)
    return {'image_urls': urls}

@app.route('/generate-all-bundle-images', methods=['POST'])
def generate_all_bundle_images():
    print("Generating All Bundle Images")
    response = requests.request("GET", ComboStoreAPI, headers={}, data={}).json()
    brands = response['brands']

    num_items = int(request.form.get('num_items'))
    introductory_discount = int(request.form.get('introductory_discount'))
    incremental_discount = int(request.form.get('incremental_discount'))
    anchor_discount = int(request.form.get('anchor_discount'))

    generate_discounts(introductory_discount, incremental_discount, anchor_discount)
    generate_quantities()

    response = requests.request("GET", ComboStoreAPI, headers={}, data={}).json()
    for dev_set in dev_sets:
        SKUs = response['SKUs'][dev_set][:num_items]
        SKUs_with_discounts = get_specific_volumes(SKUs)
        combos = generate_combos(SKUs_with_discounts, dev_set)
        parsed_json_data = {key: json.loads(value) for key, value in combos.items()}
        bundle_tables = {key: pd.read_json(StringIO(json.dumps(value)), orient='columns') for key, value in parsed_json_data.items()}
        bundle_table = pd.concat([item[1] for item in bundle_tables.items()]).set_index('BundleID')
        bundle_table = bundle_table.reset_index().rename(columns={'index': 'Bundle_ID'})
        if not os.path.exists('BundleImages'):
            os.makedirs('BundleImages')

        for index, row in bundle_table.iterrows():
            product_description = row["Product_Description"]
            dev_set = row["Dev_Set"]
            bundle_ID = row["BundleID"]
            quantities = row["Quantities"]
            max_quantity = row["Max_Quantity"]
            avg_discount = round(row["Average_Discount"], 1)

            bundle_image = generate_card_basic(
                product_description,
                dev_set,
                bundle_ID,
                quantities,
                max_quantity,
                font,
                str(avg_discount),
                theme_color=themes[brands[product_description[0]]][0],
                text_color=themes[brands[product_description[0]]][1],
                bg_color='none'
            )
            # Save the image to the SKUs directory
            bundle_image.save(f'BundleImages/{bundle_ID}.png', format='PNG')


    return "Successfully Saved All Bundle Images"

@app.route('/get_bundle_image', methods=['GET'])
def get_bundle_image():
    print("Retrieving and Resending Bundle Image from Bundle ID")
    bundle_ID = request.args.get('bundleID')
    file_name = bundle_ID + ".png"
    file_path = os.path.join(IMAGE_FOLDER, file_name)
    if os.path.exists(file_path):
        img = Image.open(file_path)
        product_cards = [(bundle_ID, img)]
        img_io = BytesIO()
        img.save(img_io, format='PNG')
        img_base64 = base64.b64encode(img_io.getvalue()).decode('utf-8')
        img_url = f"data:image/png;base64,{img_base64}"
        return jsonify({'image_url': [bundle_ID, img_url]})
    else:
        abort(404, description="File not found")

@app.route('/hello-world', methods=['GET'])
def hello_world():
    return "hello-world"

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)